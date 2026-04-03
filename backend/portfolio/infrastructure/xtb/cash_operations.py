"""
Parse XTB multi-sheet XLSX exports: ``Cash Operations`` worksheet to normalized import rows.

The worksheet has leading metadata rows; the parser locates the column header row automatically.
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, List, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from portfolio.services.transaction_import_service import NormalizedTransactionImportRow

logger = logging.getLogger(__name__)

FileArg = Union[str, Path, BytesIO, BinaryIO]

_CASH_OPERATIONS_SHEET = "cash operations"

_REQUIRED_HEADERS = frozenset(
    {"type", "ticker", "instrument", "time", "amount", "id", "comment", "product"}
)

# XTB ``Type`` values we map to portfolio buy/sell (English export).
_TYPE_BUY = frozenset(
    {
        "stock purchase",
        "purchase of shares",
        "etf purchase",
        "stock buy",
    }
)
_TYPE_SELL = frozenset(
    {
        "stock sale",
        "stock sell",
        "sale of shares",
        "etf sale",
        "stock sold",
    }
)

_COMMENT_TRADE = re.compile(
    r"(?:OPEN|CLOSE)\s+(BUY|SELL)\s+([\d.,]+)\s+@\s+([\d.,]+)",
    re.IGNORECASE,
)


def parse_xtb_cash_operations_xlsx(
    file: FileArg,
    *,
    sheet_name: Optional[str] = None,
) -> List[NormalizedTransactionImportRow]:
    """
    Read an XTB ``.xlsx`` file and return rows from the Cash Operations sheet
    that represent stock/ETF trades with parsable execution details.

    :param file: Path or binary file-like object (``bytes`` buffer seekable at 0).
    :param sheet_name: Override sheet title (default: case-insensitive ``Cash Operations``).
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    try:
        ws = _select_cash_operations_sheet(wb, sheet_name)
        header_row_idx, col_map = _find_header_row(ws)
        if header_row_idx is None or col_map is None:
            raise ValueError(
                "Could not find a header row with Type, Ticker, Time, Amount, ID, Comment."
            )
        return _parse_trade_rows(ws, header_row_idx, col_map)
    finally:
        wb.close()


def _select_cash_operations_sheet(wb, sheet_name: Optional[str]):
    target = (sheet_name or _CASH_OPERATIONS_SHEET).casefold().strip()
    for name in wb.sheetnames:
        if name.casefold().strip() == target:
            return wb[name]
    raise ValueError(
        f'Workbook has no sheet named "{sheet_name or "Cash Operations"}" '
        f"(available: {wb.sheetnames!r})"
    )


def _row_string_values(row) -> List[str]:
    return [
        "" if c.value is None else str(c.value).strip()
        for c in row
    ]


def _find_header_row(ws: Worksheet) -> Tuple[Optional[int], Optional[dict]]:
    max_scan = 80
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan), start=1):
        lowered_set = frozenset(
            str(v).strip().casefold()
            for v in _row_string_values(row)
            if str(v).strip()
        )
        if not lowered_set:
            continue
        if _REQUIRED_HEADERS.issubset(lowered_set):
            col_map = {}
            for cell in row:
                if cell.value is None:
                    continue
                key = str(cell.value).strip().casefold()
                if key in _REQUIRED_HEADERS:
                    col_map[key] = cell.column - 1
            return idx, col_map
    return None, None


def _parse_xtb_number(raw) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return float(raw)
    s = str(raw).strip().replace("\xa0", " ").replace(" ", "")
    if not s or s.lower() == "nan":
        return None
    negative = s.startswith("-")
    if negative:
        s = s[1:]
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        val = float(s)
        return -val if negative else val
    except ValueError:
        return None


def _parse_trade_datetime(raw) -> Optional[date]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_comment_trade(comment: str) -> Optional[Tuple[str, float, float]]:
    if not comment:
        return None
    m = _COMMENT_TRADE.search(comment)
    if not m:
        return None
    side = m.group(1).upper()
    qty_raw, price_raw = m.group(2), m.group(3)
    qty = _parse_xtb_number(qty_raw)
    price = _parse_xtb_number(price_raw)
    if qty is None or price is None or qty <= 0 or price <= 0:
        return None
    return side, qty, price


def _get_cell(row_values: List, col_idx: Optional[int]):
    if col_idx is None or col_idx >= len(row_values):
        return None
    return row_values[col_idx]


def _parse_trade_rows(
    ws: Worksheet,
    header_row_idx: int,
    col_map: dict,
) -> List[NormalizedTransactionImportRow]:
    out: List[NormalizedTransactionImportRow] = []
    excel_row = header_row_idx
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        excel_row += 1
        row_values = [c.value for c in row]
        type_label = _get_cell(row_values, col_map.get("type"))
        type_str = "" if type_label is None else str(type_label).strip()
        if not type_str:
            continue

        type_key = type_str.casefold()
        if type_key not in _TYPE_BUY and type_key not in _TYPE_SELL:
            continue

        ticker = _get_cell(row_values, col_map.get("ticker"))
        symbol = None if ticker is None else str(ticker).strip()
        if not symbol:
            continue

        instrument = _get_cell(row_values, col_map.get("instrument"))
        name = None if instrument is None else str(instrument).strip() or None

        time_raw = _get_cell(row_values, col_map.get("time"))
        trade_date = _parse_trade_datetime(time_raw)
        if trade_date is None:
            logger.debug("XTB import: skip row %s — bad time %r", excel_row, time_raw)
            continue

        comment_cell = _get_cell(row_values, col_map.get("comment"))
        comment = "" if comment_cell is None else str(comment_cell)

        parsed = _parse_comment_trade(comment)
        if not parsed:
            logger.debug(
                "XTB import: skip row %s — could not parse trade comment %r",
                excel_row,
                comment,
            )
            continue
        comment_side, quantity, price = parsed

        tx_side = comment_side
        if tx_side not in ("BUY", "SELL"):
            continue

        id_raw = _get_cell(row_values, col_map.get("id"))
        external_id = None if id_raw is None else str(id_raw).strip()
        if not external_id:
            continue

        out.append(
            NormalizedTransactionImportRow(
                transaction_type="BUY" if tx_side == "BUY" else "SELL",
                quantity=quantity,
                price=price,
                trade_date=trade_date,
                source_row_index=excel_row,
                symbol=symbol,
                name=name,
                asset_type="stocks",
                external_id=f"xtb:{external_id}",
                currency=None,
            )
        )
    return out
